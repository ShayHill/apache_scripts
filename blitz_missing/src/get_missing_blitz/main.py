"""Open the Blitz spreadsheet and see who has been doing their Blitzes.

:author: Shay Hill
:created: 2023-10-03
"""

from pathlib import Path
import openpyxl
from contextlib import suppress

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string

_PROJECT_DIR = Path(__file__).parent.parent.parent
_BINARIES_DIR = _PROJECT_DIR / "binaries"
_BLITZ_XLSX = _BINARIES_DIR / "2H23 Blitz - GOM DECOM.xlsx"
_PARTICIPATION_WORKSHEET = "GOM Participation"
_PARTICIPATION_HEADER_ROW = 2


def _cannonize_skip_checks(name: str) -> str:
    """Return a canonical name for a participant. Do not check aliases."""
    return name.strip().upper()


_ALIASES = {
    _cannonize_skip_checks(k): _cannonize_skip_checks(v)
    for k, v in (
        ("Gus Richmond", "Gustavus Richmond"),
        ("Derrick C. Fusilier Sr.", "Derrick Fusilier"),
    )
}


def _canonize_name(name: str) -> str:
    """Return a canonical name for a participant."""
    name = _cannonize_skip_checks(name)
    with suppress(KeyError):
        return _ALIASES[name]
    return name


def _search_value_in_col(
    ws: Worksheet, search_string: str, col: int = 1
) -> tuple[int, int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, col).value == search_string:
            return row, col
    msg = f"Could not find {search_string} in column {col}."
    raise ValueError(msg)


def _search_value_in_row(
    ws: Worksheet, search_string: str | None, row: int = 1
) -> tuple[int, int]:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row, col).value == search_string:
            return row, col
    msg = f"Could not find {search_string} in row {row}."
    raise ValueError(msg)


def _get_blitz_worksheet():
    """Get the Worksheet from the Blitz spreadsheet."""
    wb = openpyxl.load_workbook(_BLITZ_XLSX)
    return wb[_PARTICIPATION_WORKSHEET]


def _read_personnel(file: Path) -> list[str]:
    """Read list of personnel to include in report."""
    with open(file) as f:
        return [_canonize_name(x) for x in f.readlines() if not x.startswith("#")]


def _collect_bonus(ws: Worksheet):
    """Find instances where personnel participated but did not lead."""
    blitzes: dict[str, dict[str, int]] = {}
    beg_row = 2
    beg_col = column_index_from_string("AM")
    end_row = _search_value_in_col(ws, "Grand Total", beg_col)[0]
    for row in range(beg_row, end_row):
        name_value = ws.cell(row, beg_col).value
        assert isinstance(name_value, str)
        name = _canonize_name(name_value)

        bonus_value = ws.cell(row, beg_col + 1).value or 0
        assert isinstance(bonus_value, (str, int))
        bonus = int(bonus_value)
        if name in blitzes:
            blitzes[name]["Bonus"] += bonus
        else:
            blitzes[name] = {"Bonus": bonus}
    return blitzes


def _collect_blitzes(ws: Worksheet):
    """Collect the Blitzes from the spreadsheet.

    :param ws: the Worksheet from the Blitz spreadsheet
    :return: a dictionary of participants and their Blitzes

    {participant: {date_range: count, ..., 'Grand Total': count}, ...}
    """
    header_row = _PARTICIPATION_HEADER_ROW
    footer_row = _search_value_in_col(ws, "Grand Total")[0]
    right_col = _search_value_in_row(ws, None, header_row)[1]
    blitzes = _collect_bonus(ws)
    for row in range(header_row + 1, footer_row):
        participant = _canonize_name(str(ws.cell(row, 1).value))
        for col in range(2, right_col):
            date_range = str(ws.cell(header_row, col).value)
            blitzes[participant] = blitzes.get(participant) or {}
            count = ws.cell(row, col).value or 0
            assert isinstance(count, (int, str))
            blitzes[participant][date_range] = int(count)
    return blitzes


def _filter_blitzes(blitzes: dict[str, dict[str, int]], personnel: list[str]):
    """Filter blitzes to include only personnel in the personnel list."""
    filtered_blitzes = {k: {"Grand Total": 0} for k in personnel}
    filtered_blitzes.update({k: v for k, v in blitzes.items() if k in personnel})
    return filtered_blitzes


def _get_latest_blitz(blitz_line: dict[str, int]) -> str:
    """Return the latest Blitz date range."""
    dates = [k for k, v in blitz_line.items() if v > 0]
    dates = [x for x in dates if x != "Grand Total"]
    dates.insert(0, "NEVER")
    return dates[-1]


if __name__ == "__main__":
    ws = _get_blitz_worksheet()
    blitzes = _collect_blitzes(ws)
    personnel = _read_personnel(_BINARIES_DIR / "my_personnel.txt")
    filtered_blitzes = _filter_blitzes(blitzes, personnel)
    lines: list[tuple[str, str, str]] = []
    for p in sorted(filtered_blitzes, key=lambda x: x.split()[-1]):
        if filtered_blitzes[p].get("Bonus", 0) > 0:
            bonus = f"(+{filtered_blitzes[p]['Bonus']})"
        else:
            bonus = ""
        lines.append(
            (
                f"{p}",
                f"{filtered_blitzes[p]['Grand Total']}{bonus}",
                f"{_get_latest_blitz(filtered_blitzes[p])}",
            )
        )
    with (_BINARIES_DIR / "blitz_report.TXT").open("w") as f:
        _ = f.write("\n".join(["\t".join(x) for x in lines]))
    print("done")
